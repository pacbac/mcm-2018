from openpyxl import load_workbook
from openpyxl import Workbook
from globalFuncs import *
from fetchValues import indexToState, stateToIndex
#from competeResources import combineToF

sigma = {}
table = {}
msnTable = {}
proportions = {}

def combineToF(bigA, k, bigQ, sigma): #calc renewable & nonRenewable from NList
    return bigA*k**(sigma)*bigQ**(1-sigma)

#Parameters in t
def calcSigma(rConsump, rPrice , tPrice, tConsump):
    return (rConsump*rPrice)/(tConsump*tPrice)

#Parameters in t except tConsump in t+1, solve for proportion in t
def calcKnownProportions(nConsump, rConsump, tConsump, sigma):
    return tConsump/((nConsump**(sigma)*rConsump**(1-sigma)))

#Given parameters in t, calculate consumption for t+1
def calcConsumption(proportion, nConsump, rConsump, sigma):
    return proportion*nConsump**(sigma)*rConsump**(1-sigma)


def export(tbl, i):
    wb2 = Workbook()
    wb2Sheet = wb2.active

    wb2Sheet.cell(row=1, column=1).value = "Year"
    wb2Sheet.cell(row=1, column=2).value = "A(Year)"

    r = 2
    for yr in tbl.keys():
        wb2Sheet.cell(row=r, column=1).value = yr
        wb2Sheet.cell(row=r, column=2).value = tbl[yr]
        r += 1
    wb2.save("totalAvgPrice"+indexToState(i)+".xlsx")
    pass

def load():
    wb1 = load_workbook("ProblemCData.xlsx")
    dataSheet = wb1.worksheets[0]
    msnCodes = wb1.worksheets[1]

    global table
    global msnTable
    global sigma
    global proportions
    for r in range(2, dataSheet.max_row+1):
        msn = dataSheet.cell(row=r, column=1).value #col 1
        state = stateToIndex(dataSheet.cell(row=r, column=2).value) #col 2
        yr = int(dataSheet.cell(row=r, column=3).value) #col 3
        data = float(dataSheet.cell(row=r, column=4).value) #col 4
        if msn not in table:
            table[msn] = [{}, {}, {}, {}] #0=AZ, 1=CA, 2=NM, 3=TX
        table[msn][state][yr] = data

    for r in range(2, msnCodes.max_row+1):
        msn = msnCodes.cell(row=r, column=1).value
        desc = msnCodes.cell(row=r, column=2).value
        unit = msnCodes.cell(row=r, column=3).value
        msnTable[msn] = [desc, unit] #array of [description, unit of msn]

    for i in range(0, 4):
        #total consumption
        tc = {yr: table["TETCB"][i][yr] for yr in table["GDPRX"][i].keys()}
        #total average price
        tap = {yr: table["TETCD"][i][yr] for yr in table["GDPRX"][i].keys()}
        #total non renewable average price
        tnrap = {yr: (table["PATCD"][i][yr]+table["NGTCD"][i][yr])/2 for yr in table["GDPRX"][i].keys()}
        #total non renewable consumption
        tnrc = {yr: (table["PATCB"][i][yr]+table["NGTCB"][i][yr]) for yr in table["GDPRX"][i].keys()}
        #total renewable consumption
        trc = {yr: table["RETCB"][i][yr] for yr in table["GDPRX"][i].keys()}
        #total renewable average price
        trap = {yr: float((tc[yr]*tap[yr]-tnrc[yr]*tnrap[yr])/trc[yr]) for yr in table["GDPRX"][i].keys()}

        for yr in table["GDPRX"][i].keys():
            sigma[yr] = calcSigma(trc[yr], trap[yr], tc[yr], tap[yr]) #k, q, p, f
        for j in range(1978, 2010):
            proportions[j-1] = calcKnownProportions(tnrc[j-1], trc[j-1], tc[j], sigma[j-1]) #k, q, f, sigma

        if __name__ == "__main__":
            export(tap, i)
    pass

load()
